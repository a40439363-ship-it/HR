from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "Yuborilgan vaqt",
    "Nomzod username",
    "F.I.Sh",
    "Tug'ilgan sana",
    "Sertifikat",
    "Telefon raqam",
    "Filial",
    "Ariza bo'limi",
    "Yo'nalish / Lavozim",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws) -> None:
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = max(18, len(header) + 4)
    ws.freeze_panes = "A2"


def ensure_workbook(file_path: str) -> None:
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Arizalar"
        ws.append(HEADERS)
        _style_header(ws)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.title = "Arizalar"

    current_headers = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    if current_headers != HEADERS:
        rows_to_keep = []
        if ws.max_row >= 2:
            existing_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            header_map = {str(h).strip(): idx for idx, h in enumerate(existing_headers, start=1) if h}
            synonyms = {
                "Filial": ["Filial"],
                "Yuborilgan vaqt": ["Yuborilgan vaqt"],
                "Nomzod username": ["Nomzod username"],
                "F.I.Sh": ["F.I.Sh"],
                "Tug'ilgan sana": ["Tug'ilgan sana"],
                "Sertifikat": ["Sertifikat"],
                "Telefon raqam": ["Telefon raqam"],
                "Ariza bo'limi": ["Ariza bo'limi"],
                "Yo'nalish / Lavozim": ["Yo'nalish / Lavozim"],
            }
            for row_idx in range(2, ws.max_row + 1):
                row = []
                for header in HEADERS:
                    col_idx = None
                    for name in synonyms.get(header, [header]):
                        col_idx = header_map.get(name)
                        if col_idx:
                            break
                    row.append(ws.cell(row=row_idx, column=col_idx).value if col_idx else "")
                if any(v not in (None, "") for v in row):
                    rows_to_keep.append(row)

        ws.delete_rows(1, ws.max_row)
        ws.append(HEADERS)
        for row in rows_to_keep:
            ws.append(row)
        _style_header(ws)
        wb.save(path)
    else:
        _style_header(ws)
        wb.save(path)


def append_application(file_path: str, row_data: list) -> None:
    ensure_workbook(file_path)
    wb = load_workbook(file_path)
    ws = wb["Arizalar"]
    ws.append(row_data)
    wb.save(file_path)
